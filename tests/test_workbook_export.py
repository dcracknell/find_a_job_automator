from __future__ import annotations

import sqlite3
from datetime import date, timedelta

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from job_search.output.workbook_export import _infer_experience_level, regenerate_workbook
from job_search.output.workbook_import import import_user_edits
from job_search.storage.db import migrate


def _conn_with_job() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    migrate(conn=conn)
    conn.execute(
        """
        INSERT INTO jobs (
            job_id, source, matched_query, first_seen, last_seen, status,
            title, company, location, url, description, posted_date, closes_on,
            salary_raw, salary_min, salary_max, fit_score, fit_confidence, fit_reason,
            matched_keywords, ranker_version, jd_content_hash
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "job-1",
            "adzuna",
            "python developer sheffield",
            date.today().isoformat(),
            date.today().isoformat(),
            "new",
            "Graduate Python Developer",
            "Example Ltd",
            "Sheffield",
            "https://example.com/jobs/1",
            "Build useful services in Python.",
            date.today().isoformat(),
            (date.today() + timedelta(days=6)).isoformat(),
            "GBP 30000-40000",
            30000,
            40000,
            8.7,
            0.82,
            "Python and service work align closely with the profile.",
            '["Python", "services"]',
            "v1-test",
            "hash-1",
        ),
    )
    conn.commit()
    return conn


def _headers(ws) -> list[str]:
    return [cell.value for cell in ws[1]]


def test_workbook_export_adds_readable_tracking_columns(tmp_path) -> None:
    conn = _conn_with_job()
    xlsx_path = tmp_path / "jobs.xlsx"
    regenerate_workbook(conn, xlsx_path=xlsx_path, backups_dir=tmp_path / "backups")

    wb = load_workbook(xlsx_path, data_only=False)
    try:
        assert wb.sheetnames[:2] == ["overview", "jobs"]

        ws = wb["jobs"]
        headers = _headers(ws)
        assert headers == [
            "Job ID",
            "Status",
            "Priority",
            "Next Step",
            "Score",
            "Confidence",
            "Score Band",
            "Experience Level",
            "Title",
            "Company",
            "Location",
            "Salary (raw)",
            "Salary Min",
            "Salary Max",
            "Posted",
            "Closes",
            "Days Left",
            "Source",
            "Apply Link",
            "Anthropic Response",
            "Matched Keywords",
            "First Seen",
            "Last Seen",
            "Query",
            "Notes",
            "Ranker Ver.",
        ]

        col = {header: idx + 1 for idx, header in enumerate(headers)}
        assert ws.cell(2, col["Priority"]).value.startswith("=IF(")
        assert ws.cell(2, col["Next Step"]).value.startswith("=IF(")
        assert ws.cell(2, col["Score Band"]).value.startswith("=IF(")
        assert ws.cell(2, col["Experience Level"]).value == "Graduate/Entry"
        assert ws.cell(2, col["Days Left"]).value.startswith("=IF(")
        assert ws.cell(2, col["Anthropic Response"]).value == (
            "Python and service work align closely with the profile."
        )
        assert ws.cell(2, col["Matched Keywords"]).value == "Python, services"

        link_cell = ws.cell(2, col["Apply Link"])
        assert link_cell.value == "Open posting"
        assert link_cell.hyperlink.target == "https://example.com/jobs/1"

        assert ws.freeze_panes == "I2"
        assert ws.sheet_view.zoomScale == 90
        for header in (
            "Job ID",
            "Confidence",
            "Salary Min",
            "Salary Max",
            "Source",
            "Anthropic Response",
            "Matched Keywords",
            "First Seen",
            "Last Seen",
            "Query",
            "Ranker Ver.",
        ):
            assert ws.column_dimensions[get_column_letter(col[header])].hidden is True

        validations = ws.data_validations.dataValidation
        assert len(validations) == 1
        assert validations[0].formula1 == (
            '"new,applied,interview,offer,rejected,ignore,archive,closed"'
        )
        assert ws.cell(1, col["Status"]).comment is not None
        assert ws.cell(1, col["Notes"]).comment is not None
    finally:
        wb.close()
        conn.close()


def test_exported_workbook_still_imports_status_and_notes(tmp_path) -> None:
    conn = _conn_with_job()
    xlsx_path = tmp_path / "jobs.xlsx"
    regenerate_workbook(conn, xlsx_path=xlsx_path, backups_dir=tmp_path / "backups")

    wb = load_workbook(xlsx_path)
    try:
        ws = wb["jobs"]
        headers = _headers(ws)
        col = {header: idx + 1 for idx, header in enumerate(headers)}
        ws.cell(2, col["Status"]).value = "applied"
        ws.cell(2, col["Notes"]).value = "Applied 2026-06-17; follow up next week."
        wb.save(xlsx_path)
    finally:
        wb.close()

    assert import_user_edits(conn, xlsx_path=xlsx_path) == 1
    row = conn.execute("SELECT status, notes FROM jobs WHERE job_id = 'job-1'").fetchone()
    assert row["status"] == "applied"
    assert row["notes"] == "Applied 2026-06-17; follow up next week."
    conn.close()


def test_infer_experience_level_from_title_and_description() -> None:
    assert _infer_experience_level("2026 Graduate Firmware Engineer") == "Graduate/Entry"
    assert _infer_experience_level("Senior Software Engineer") == "Senior/Lead"
    assert _infer_experience_level("Project Manager") == "Manager/Owner"
    assert _infer_experience_level("Software Engineer", "Requires 3+ years Python.") == "Mid-level"
