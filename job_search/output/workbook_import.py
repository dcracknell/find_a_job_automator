"""Import user edits from jobs.xlsx back into SQLite.

Called at the start of each run, before the pipeline touches the DB.
Only imports 'status' and 'notes' — the two columns users are expected to edit.
Uses file mtime vs DB last_user_edit for conflict detection.
On true conflict (both sides changed), Excel wins with a logged warning.
"""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime
from pathlib import Path

from job_search import PROJECT_ROOT

logger = logging.getLogger(__name__)

_XLSX_PATH: Path = PROJECT_ROOT / "data" / "jobs.xlsx"

# Valid status values — anything else is rejected with a warning
_VALID_STATUSES = frozenset(
    ["new", "applied", "interview", "offer", "rejected", "ignore", "archive", "closed"]
)


def import_user_edits(
    conn: sqlite3.Connection,
    xlsx_path: Path | None = None,
) -> int:
    """Read user-edited status and notes from xlsx and write changes to SQLite.

    Returns the number of rows updated.
    If xlsx_path doesn't exist or is unreadable, logs a warning and returns 0.
    """
    xlsx_path = xlsx_path or _XLSX_PATH

    if not xlsx_path.exists():
        logger.debug("import_user_edits: %s not found, skipping", xlsx_path)
        return 0

    # Get the file's mtime as an ISO timestamp
    try:
        mtime = datetime.utcfromtimestamp(xlsx_path.stat().st_mtime)
        mtime_str = mtime.isoformat() + "Z"
    except OSError as exc:
        logger.warning("import_user_edits: could not stat %s: %s", xlsx_path, exc)
        return 0

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("import_user_edits: could not open %s: %s — skipping", xlsx_path, exc)
        return 0

    ws = wb["jobs"] if "jobs" in wb.sheetnames else None
    if ws is None:
        logger.warning("import_user_edits: 'jobs' sheet not found in %s", xlsx_path)
        wb.close()
        return 0

    # Read header row to find column indices
    rows = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        wb.close()
        return 0

    try:
        job_id_idx = headers.index("Job ID")
        status_idx = headers.index("Status")
        notes_idx = headers.index("Notes")
    except ValueError as exc:
        logger.warning("import_user_edits: missing column in %s: %s", xlsx_path, exc)
        wb.close()
        return 0

    updated = 0
    conflicts = 0

    for row in rows:
        if row[job_id_idx] is None:
            continue

        job_id = str(row[job_id_idx]).strip()
        new_status = str(row[status_idx]).strip() if row[status_idx] is not None else ""
        new_notes = str(row[notes_idx]).strip() if row[notes_idx] is not None else ""

        if not job_id:
            continue

        # Validate status value
        if new_status and new_status not in _VALID_STATUSES:
            logger.warning(
                "import_user_edits: invalid status '%s' for job %s — skipping row",
                new_status,
                job_id,
            )
            continue

        # Fetch existing DB row
        db_row = conn.execute(
            "SELECT status, notes, last_user_edit FROM jobs WHERE job_id = ?",
            (job_id,),
        ).fetchone()

        if db_row is None:
            continue  # job no longer in DB; skip

        db_status = db_row["status"] or ""
        db_notes = db_row["notes"] or ""
        db_last_edit = db_row["last_user_edit"]  # may be None

        # Skip if nothing changed
        if new_status == db_status and new_notes == db_notes:
            continue

        # Conflict check: if DB was edited after the xlsx was last modified
        if db_last_edit is not None:
            try:
                db_edit_dt = datetime.fromisoformat(db_last_edit.rstrip("Z"))
                if db_edit_dt > mtime:
                    # DB was updated more recently than the xlsx — true conflict
                    conflicts += 1
                    logger.warning(
                        "import_user_edits: conflict on %s — DB edited %s, xlsx mtime %s. "
                        "Excel wins.",
                        job_id,
                        db_edit_dt.isoformat(),
                        mtime.isoformat(),
                    )
            except (ValueError, AttributeError):
                pass  # unparseable timestamp; treat as no conflict

        # Apply Excel's values (Excel wins, as per spec)
        conn.execute(
            "UPDATE jobs SET status = ?, notes = ?, last_user_edit = ? WHERE job_id = ?",
            (new_status or db_status, new_notes, mtime_str, job_id),
        )
        updated += 1

    conn.commit()
    wb.close()

    if updated:
        logger.info("import_user_edits: updated %d row(s) from %s", updated, xlsx_path)
    if conflicts:
        logger.warning(
            "import_user_edits: %d conflict(s) detected (Excel won in all cases)", conflicts
        )

    return updated
