"""Regenerate data/jobs.xlsx from SQLite.

Writes are always atomic (build as .tmp, rename on success) and always back up
any existing file first. The atomic_xlsx_write helper is reusable by other modules.
"""

from __future__ import annotations

import json
import logging
import os
import shutil
import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from job_search import PROJECT_ROOT

logger = logging.getLogger(__name__)

_XLSX_PATH: Path = PROJECT_ROOT / "data" / "jobs.xlsx"
_BACKUPS_DIR: Path = PROJECT_ROOT / "data" / "backups"

# Column definitions for the 'jobs' sheet: (header, db_column).
# Columns with db_column=None are workbook-only formula/helper columns.
_JOBS_COLUMNS: list[tuple[str, str | None]] = [
    ("Job ID", "job_id"),
    ("Status", "status"),
    ("Priority", None),
    ("Next Step", None),
    ("Score", "fit_score"),
    ("Confidence", "fit_confidence"),
    ("Score Band", None),
    ("Title", "title"),
    ("Company", "company"),
    ("Location", "location"),
    ("Salary (raw)", "salary_raw"),
    ("Salary Min", "salary_min"),
    ("Salary Max", "salary_max"),
    ("Posted", "posted_date"),
    ("Closes", "closes_on"),
    ("Days Left", None),
    ("Source", "source"),
    ("Apply Link", "url"),
    ("Claude's Fit Reason", "fit_reason"),
    ("Matched Keywords", "matched_keywords"),
    ("First Seen", "first_seen"),
    ("Last Seen", "last_seen"),
    ("Query", "matched_query"),
    ("Notes", "notes"),
    ("Ranker Ver.", "ranker_version"),
]

# Status values users can set in Excel and import back into SQLite.
_VALID_STATUSES: tuple[str, ...] = (
    "new",
    "applied",
    "interview",
    "offer",
    "rejected",
    "ignore",
    "archive",
    "closed",
)

# Status values and their row fill colours.
_STATUS_FILLS: dict[str, str] = {
    "new": "EAF3F8",
    "applied": "D9EAD3",
    "interview": "D9E2F3",
    "offer": "B6D7A8",
    "rejected": "F4CCCC",
    "archive": "D3D3D3",  # light grey
    "closed": "BEBEBE",   # grey
    "ignore": "F0E68C",   # khaki
}

_PRIORITY_FILLS: dict[str, str] = {
    "P1 Apply": "F4CCCC",
    "P2 Strong": "FCE5CD",
    "P3 Maybe": "FFF2CC",
    "P4 Low": "EADCF8",
    "Applied": "D9EAD3",
    "Interview": "D9E2F3",
    "Offer": "B6D7A8",
    "Rejected": "E6B8AF",
    "Closed": "D9D9D9",
    "Ignore": "D9D9D9",
    "Archive": "D9D9D9",
    "Unscored": "E7E6E6",
}

_SCORE_GREEN = "00B050"   # fit_score ≥ 8
_SCORE_AMBER = "FFC000"   # fit_score 5–7.9
_SCORE_RED = "FF0000"     # fit_score < 5
_CLOSES_SOON = "FF0000"   # closes within 7 days

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
_FORMULA_FILL = PatternFill("solid", fgColor="EAF3F8")
_LIGHT_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))
_DATE_HEADERS = {"Posted", "Closes", "First Seen", "Last Seen"}
_MONEY_HEADERS = {"Salary Min", "Salary Max"}
_WRAPPED_HEADERS = {
    "Title",
    "Location",
    "Salary (raw)",
    "Claude's Fit Reason",
    "Matched Keywords",
    "Query",
    "Notes",
}


def atomic_xlsx_write(workbook: Workbook, dest_path: Path) -> None:
    """Save a workbook atomically: write to .tmp, then rename to dest.

    This is the ONLY way workbooks should be saved anywhere in the codebase.
    Never call workbook.save(final_path) directly.
    """
    tmp_path = dest_path.with_suffix(".xlsx.tmp")
    try:
        workbook.save(str(tmp_path))
        # On Windows, os.replace is atomic within the same volume
        os.replace(str(tmp_path), str(dest_path))
        logger.debug("Atomically wrote %s", dest_path)
    except Exception:
        if tmp_path.exists():
            tmp_path.unlink()
        raise


def backup_existing_xlsx(xlsx_path: Path, backups_dir: Path) -> None:
    """Copy xlsx_path to backups_dir/jobs.YYYY-MM-DD.xlsx (no-op if file missing)."""
    if not xlsx_path.exists():
        return
    backups_dir.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    dest = backups_dir / f"jobs.{today}.xlsx"
    for i in range(1, 1000):
        if not dest.exists():
            break
        dest = backups_dir / f"jobs.{today}.{i}.xlsx"
    else:
        raise RuntimeError(f"Could not find available backup path in {backups_dir}")

    shutil.copy2(str(xlsx_path), str(dest))
    logger.info("Backed up existing workbook to %s", dest)


def _db_columns() -> list[str]:
    """Return unique DB columns required for the jobs workbook."""
    seen: set[str] = set()
    cols: list[str] = []
    for _, db_col in _JOBS_COLUMNS:
        if db_col and db_col not in seen:
            seen.add(db_col)
            cols.append(db_col)
    return cols


def _header_map(headers: list[str]) -> dict[str, int]:
    """Map header name to 1-based worksheet column index."""
    return {header: i for i, header in enumerate(headers, start=1)}


def _cell_ref(headers: list[str], header: str, row_idx: int) -> str:
    col = get_column_letter(_header_map(headers)[header])
    return f"${col}{row_idx}"


def _parse_iso_date(value: object) -> date | None:
    """Convert SQLite ISO date strings into date objects for Excel date formulas."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _format_keywords(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    text = str(value)
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return text
    if isinstance(parsed, list):
        return ", ".join(str(v) for v in parsed)
    return text


def _format_job_value(header: str, value: object) -> object:
    if header in _DATE_HEADERS:
        return _parse_iso_date(value)
    if header == "Matched Keywords":
        return _format_keywords(value)
    return value


def _priority_label(status: str, score: float | None, closes_on: object) -> str:
    status = (status or "new").lower()
    status_labels = {
        "applied": "Applied",
        "interview": "Interview",
        "offer": "Offer",
        "rejected": "Rejected",
        "ignore": "Ignore",
        "archive": "Archive",
        "closed": "Closed",
    }
    if status in status_labels:
        return status_labels[status]
    if score is None:
        return "Unscored"

    close_date = _parse_iso_date(closes_on)
    closes_soon = close_date is None or (close_date - date.today()).days <= 14
    if score >= 8 and closes_soon:
        return "P1 Apply"
    if score >= 7:
        return "P2 Strong"
    if score >= 5:
        return "P3 Maybe"
    return "P4 Low"


def _sort_key_for_priority(label: str) -> tuple[int, str]:
    order = {
        "P1 Apply": 0,
        "P2 Strong": 1,
        "Interview": 2,
        "Applied": 3,
        "Offer": 4,
        "P3 Maybe": 5,
        "P4 Low": 6,
        "Unscored": 7,
        "Rejected": 8,
        "Closed": 9,
        "Ignore": 10,
        "Archive": 11,
    }
    return (order.get(label, 99), label)


def _excel_if(condition: str, true_value: str, false_value: str) -> str:
    return f"IF({condition},{true_value},{false_value})"


def _quoted(value: str) -> str:
    return f'"{value}"'


def _priority_formula(headers: list[str], row_idx: int) -> str:
    status = _cell_ref(headers, "Status", row_idx)
    score = _cell_ref(headers, "Score", row_idx)
    closes = _cell_ref(headers, "Closes", row_idx)

    formula = _quoted("P4 Low")
    formula = _excel_if(f"{score}>=5", _quoted("P3 Maybe"), formula)
    formula = _excel_if(f"{score}>=7", _quoted("P2 Strong"), formula)
    formula = _excel_if(
        f"AND({score}>=8,OR({closes}=\"\",{closes}-TODAY()<=14))",
        _quoted("P1 Apply"),
        formula,
    )
    formula = _excel_if(f"{score}=\"\"", _quoted("Unscored"), formula)

    status_labels = (
        ("closed", "Closed"),
        ("archive", "Archive"),
        ("ignore", "Ignore"),
        ("rejected", "Rejected"),
        ("offer", "Offer"),
        ("interview", "Interview"),
        ("applied", "Applied"),
    )
    for status_value, label in status_labels:
        formula = _excel_if(f"{status}=\"{status_value}\"", _quoted(label), formula)
    return f"={formula}"


def _next_step_formula(headers: list[str], row_idx: int) -> str:
    status = _cell_ref(headers, "Status", row_idx)
    priority = _cell_ref(headers, "Priority", row_idx)

    formula = _quoted("Skip unless useful")
    formula = _excel_if(f"{priority}=\"P3 Maybe\"", _quoted("Save if interesting"), formula)
    formula = _excel_if(f"{priority}=\"P2 Strong\"", _quoted("Review next"), formula)
    formula = _excel_if(f"{priority}=\"P1 Apply\"", _quoted("Apply today"), formula)

    status_actions = (
        ("closed", "Closed"),
        ("archive", "No action"),
        ("ignore", "No action"),
        ("rejected", "Archive lessons"),
        ("offer", "Compare offer"),
        ("interview", "Prep interview"),
        ("applied", "Follow up"),
    )
    for status_value, action in status_actions:
        formula = _excel_if(f"{status}=\"{status_value}\"", _quoted(action), formula)
    return f"={formula}"


def _score_band_formula(headers: list[str], row_idx: int) -> str:
    score = _cell_ref(headers, "Score", row_idx)
    formula = _quoted("Low 0-2")
    formula = _excel_if(f"{score}>=3", _quoted("Weak 3-4"), formula)
    formula = _excel_if(f"{score}>=5", _quoted("Possible 5-6"), formula)
    formula = _excel_if(f"{score}>=7", _quoted("Strong 7-8"), formula)
    formula = _excel_if(f"{score}>=9", _quoted("Excellent 9-10"), formula)
    formula = _excel_if(f"{score}=\"\"", _quoted("Unscored"), formula)
    return f"={formula}"


def _days_left_formula(headers: list[str], row_idx: int) -> str:
    closes = _cell_ref(headers, "Closes", row_idx)
    return f'=IF({closes}="","",{closes}-TODAY())'


def _apply_job_formulas(ws, headers: list[str], row_idx: int) -> None:
    cols = _header_map(headers)
    ws.cell(row=row_idx, column=cols["Priority"]).value = _priority_formula(headers, row_idx)
    ws.cell(row=row_idx, column=cols["Next Step"]).value = _next_step_formula(headers, row_idx)
    ws.cell(row=row_idx, column=cols["Score Band"]).value = _score_band_formula(headers, row_idx)
    ws.cell(row=row_idx, column=cols["Days Left"]).value = _days_left_formula(headers, row_idx)


def _style_header_row(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _LIGHT_BORDER


def _add_jobs_table(ws) -> None:
    if ws.max_row <= 1:
        return
    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="JobsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _add_status_dropdown(ws, headers: list[str]) -> None:
    status_col = get_column_letter(_header_map(headers)["Status"])
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(_VALID_STATUSES)}"',
        allow_blank=False,
    )
    validation.error = "Choose one of: " + ", ".join(_VALID_STATUSES)
    validation.errorTitle = "Invalid status"
    validation.prompt = "Use applied/interview/offer/rejected to track applications."
    validation.promptTitle = "Track this job"
    ws.add_data_validation(validation)
    validation.add(f"{status_col}2:{status_col}{max(ws.max_row + 100, 1000)}")


def _style_jobs_sheet(ws, headers: list[str]) -> None:
    cols = _header_map(headers)
    ws.freeze_panes = "E2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    _style_header_row(ws)

    col_widths = {
        "Job ID": 10,
        "Status": 13,
        "Priority": 13,
        "Next Step": 16,
        "Score": 8,
        "Confidence": 11,
        "Score Band": 16,
        "Title": 38,
        "Company": 24,
        "Location": 24,
        "Salary (raw)": 20,
        "Salary Min": 12,
        "Salary Max": 12,
        "Posted": 12,
        "Closes": 12,
        "Days Left": 10,
        "Source": 16,
        "Apply Link": 16,
        "Claude's Fit Reason": 52,
        "Matched Keywords": 30,
        "First Seen": 12,
        "Last Seen": 12,
        "Query": 28,
        "Notes": 48,
        "Ranker Ver.": 12,
    }
    for header in headers:
        letter = get_column_letter(cols[header])
        ws.column_dimensions[letter].width = col_widths.get(header, 15)

    for header in ("Job ID", "Ranker Ver."):
        ws.column_dimensions[get_column_letter(cols[header])].hidden = True

    for header in ("Status", "Notes"):
        cell = ws.cell(row=1, column=cols[header])
        cell.fill = _EDITABLE_FILL
        cell.font = Font(bold=True, color="000000")

    ws.cell(row=1, column=cols["Status"]).comment = Comment(
        "Pick from the dropdown. Use 'applied' after submitting an application.",
        "job-search",
    )
    ws.cell(row=1, column=cols["Notes"]).comment = Comment(
        "Add application date, contact name, follow-up reminder, or anything you want to keep.",
        "job-search",
    )
    ws.cell(row=1, column=cols["Claude's Fit Reason"]).comment = Comment(
        "This is the model's fit rationale from the ranking step.",
        "job-search",
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 48
        for cell in row:
            header = headers[cell.column - 1]
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=header in _WRAPPED_HEADERS,
            )
            if header in ("Status", "Notes"):
                cell.fill = _EDITABLE_FILL
            elif header in ("Priority", "Next Step", "Score Band", "Days Left"):
                cell.fill = _FORMULA_FILL

    for header in _DATE_HEADERS:
        col_idx = cols[header]
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for c in cell:
                c.number_format = "d mmm yyyy"

    for header in _MONEY_HEADERS:
        col_idx = cols[header]
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for c in cell:
                c.number_format = '£#,##0'

    for header, number_format in {"Score": "0.0", "Confidence": "0%", "Days Left": "0"}.items():
        col_idx = cols[header]
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for c in cell:
                c.number_format = number_format

    _add_status_dropdown(ws, headers)
    _add_jobs_table(ws)


def _add_jobs_conditional_formatting(ws, headers: list[str]) -> None:
    last_row = ws.max_row
    if last_row <= 1:
        return

    cols = _header_map(headers)
    score_col = get_column_letter(cols["Score"])
    score_range = f"{score_col}2:{score_col}{last_row}"

    green_fill = PatternFill(start_color=_SCORE_GREEN, end_color=_SCORE_GREEN, fill_type="solid")
    ws.conditional_formatting.add(
        score_range,
        FormulaRule(formula=[f"{score_col}2>=8"], fill=green_fill),
    )

    amber_fill = PatternFill(start_color=_SCORE_AMBER, end_color=_SCORE_AMBER, fill_type="solid")
    ws.conditional_formatting.add(
        score_range,
        FormulaRule(formula=[f"AND({score_col}2>=5,{score_col}2<8)"], fill=amber_fill),
    )

    red_fill = PatternFill(start_color=_SCORE_RED, end_color=_SCORE_RED, fill_type="solid")
    ws.conditional_formatting.add(
        score_range,
        FormulaRule(formula=[f'AND({score_col}2<>"",{score_col}2<5)'], fill=red_fill),
    )

    status_col = get_column_letter(cols["Status"])
    status_range = f"{status_col}2:{status_col}{last_row}"
    for status, colour in _STATUS_FILLS.items():
        fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[f'{status_col}2="{status}"'], fill=fill),
        )

    priority_col = get_column_letter(cols["Priority"])
    priority_range = f"{priority_col}2:{priority_col}{last_row}"
    for label, colour in _PRIORITY_FILLS.items():
        fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        ws.conditional_formatting.add(
            priority_range,
            FormulaRule(formula=[f'{priority_col}2="{label}"'], fill=fill),
        )

    days_col = get_column_letter(cols["Days Left"])
    days_range = f"{days_col}2:{days_col}{last_row}"
    soon_font = Font(bold=True, color=_SCORE_RED)
    soon_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws.conditional_formatting.add(
        days_range,
        FormulaRule(
            formula=[f'AND({days_col}2<>"",{days_col}2>=0,{days_col}2<=7)'],
            font=soon_font,
            fill=soon_fill,
        ),
    )


def _build_overview_sheet(ws, conn: sqlite3.Connection) -> None:
    """Populate a quick-read summary sheet for priorities and application tracking."""
    ws.title = "overview"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    title = ws["A1"]
    title.value = "Job Search Overview"
    title.font = Font(bold=True, size=18, color="1F4E78")
    ws["A2"] = "Last exported"
    ws["B2"] = datetime.now().replace(microsecond=0)
    ws["B2"].number_format = "d mmm yyyy h:mm"

    ws["A4"] = (
        "Use Status to track each role. Set Status to applied when you apply, "
        "then add the date and details in Notes."
    )
    ws["A4"].alignment = Alignment(wrap_text=True)

    ws["A6"] = "Status"
    ws["B6"] = "Count"
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.fill = _SUBHEADER_FILL
    for row_idx, status in enumerate(_VALID_STATUSES, start=7):
        ws.cell(row=row_idx, column=1).value = status
        ws.cell(row=row_idx, column=2).value = f'=COUNTIF(jobs!$B:$B,A{row_idx})'

    priority_start = 17
    ws.cell(row=priority_start, column=1).value = "Priority"
    ws.cell(row=priority_start, column=2).value = "Count"
    for cell in ws[priority_start]:
        cell.font = Font(bold=True)
        cell.fill = _SUBHEADER_FILL
    priority_labels = (
        "P1 Apply",
        "P2 Strong",
        "P3 Maybe",
        "P4 Low",
        "Applied",
        "Interview",
        "Offer",
    )
    for row_idx, label in enumerate(priority_labels, start=priority_start + 1):
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=2).value = f'=COUNTIF(jobs!$C:$C,A{row_idx})'

    top_start = 27
    ws.cell(row=top_start, column=1).value = "Top Priorities"
    ws.cell(row=top_start, column=1).font = Font(bold=True, size=14, color="1F4E78")

    headers = [
        "Priority",
        "Score",
        "Status",
        "Title",
        "Company",
        "Location",
        "Closes",
        "Apply Link",
        "Claude's Fit Reason",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=top_start + 1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    rows = conn.execute(
        """
        SELECT status, fit_score, title, company, location, closes_on, url, fit_reason
        FROM jobs
        ORDER BY fit_score DESC NULLS LAST, closes_on IS NULL, closes_on ASC, first_seen DESC
        LIMIT 50
        """
    ).fetchall()
    scored_rows = []
    for row in rows:
        priority = _priority_label(row["status"], row["fit_score"], row["closes_on"])
        scored_rows.append((priority, row))
    scored_rows.sort(
        key=lambda item: (
            _sort_key_for_priority(item[0]),
            -(item[1]["fit_score"] or -1),
            str(item[1]["closes_on"] or "9999-12-31"),
        )
    )

    for row_idx, (priority, row) in enumerate(scored_rows[:15], start=top_start + 2):
        values = [
            priority,
            row["fit_score"],
            row["status"],
            row["title"],
            row["company"],
            row["location"],
            _parse_iso_date(row["closes_on"]),
            "Open posting" if row["url"] else "",
            row["fit_reason"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=col_idx in (4, 6, 9))
        if row["url"]:
            link_cell = ws.cell(row=row_idx, column=8)
            link_cell.hyperlink = row["url"]
            link_cell.style = "Hyperlink"

    widths = {
        "A": 16,
        "B": 10,
        "C": 12,
        "D": 38,
        "E": 24,
        "F": 24,
        "G": 12,
        "H": 16,
        "I": 52,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    ws.row_dimensions[4].height = 34


def _build_jobs_sheet(ws, conn: sqlite3.Connection) -> None:
    """Populate the 'jobs' worksheet from SQLite."""
    headers = [h for h, _ in _JOBS_COLUMNS]
    db_cols = _db_columns()

    ws.append(headers)

    rows = conn.execute(
        f"SELECT {', '.join(db_cols)} FROM jobs ORDER BY fit_score DESC NULLS LAST, first_seen DESC"
    ).fetchall()

    cols = _header_map(headers)
    for row_idx, row in enumerate(rows, start=2):
        values = [
            _format_job_value(header, row[db_col]) if db_col else None
            for header, db_col in _JOBS_COLUMNS
        ]
        ws.append(values)
        _apply_job_formulas(ws, headers, row_idx)
        if row["url"]:
            link_cell = ws.cell(row=row_idx, column=cols["Apply Link"])
            link_cell.value = "Open posting"
            link_cell.hyperlink = row["url"]
            link_cell.style = "Hyperlink"

    _style_jobs_sheet(ws, headers)
    _add_jobs_conditional_formatting(ws, headers)


def _build_runs_sheet(ws, conn: sqlite3.Connection) -> None:
    """Populate the 'runs' worksheet with the last 30 runs."""
    headers = ["ID", "Started", "Finished", "Duration (s)",
               "Sources OK", "Sources Failed", "Jobs Scraped",
               "Jobs New", "Jobs Closed", "Errors"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = conn.execute(
        "SELECT id, started_at, finished_at, duration_s, sources_ok, sources_failed, "
        "jobs_scraped, jobs_new, jobs_closed, errors FROM runs ORDER BY id DESC LIMIT 30"
    ).fetchall()
    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"


def _build_profile_sheet(ws) -> None:
    """Populate the 'profile' worksheet as a read-only mirror of profile.json."""
    import json
    profile_path = PROJECT_ROOT / "config" / "profile.json"
    ws.append(["Key", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    if not profile_path.exists():
        ws.append(["(profile.json not found)", ""])
        return

    try:
        with profile_path.open() as f:
            profile = json.load(f)
    except Exception as exc:
        ws.append(["(error reading profile.json)", str(exc)])
        return

    def _flatten(d: dict, prefix: str = "") -> list[tuple[str, str]]:
        items = []
        for k, v in d.items():
            key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                items.extend(_flatten(v, key))
            else:
                items.append((key, json.dumps(v) if not isinstance(v, str) else v))
        return items

    for k, v in _flatten(profile):
        ws.append([k, v])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


def _build_quota_sheet(ws, conn: sqlite3.Connection) -> None:
    """Populate the 'quota' worksheet with last 30 days of API spend."""
    headers = ["Date", "Operation", "Model", "Calls",
               "Input Tokens", "Cached Tokens", "Output Tokens", "Est. Cost (GBP)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = conn.execute(
        """
        SELECT
            DATE(timestamp) as day,
            operation,
            model,
            COUNT(*) as calls,
            SUM(input_tokens) as input_tokens,
            SUM(cached_input_tokens) as cached_tokens,
            SUM(output_tokens) as output_tokens,
            ROUND(SUM(est_cost_gbp), 4) as est_cost_gbp
        FROM api_calls
        WHERE DATE(timestamp) >= DATE('now', '-30 days')
        GROUP BY day, operation, model
        ORDER BY day DESC, est_cost_gbp DESC
        """
    ).fetchall()
    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"


def regenerate_workbook(
    conn: sqlite3.Connection,
    xlsx_path: Path | None = None,
    backups_dir: Path | None = None,
) -> None:
    """Regenerate the Excel workbook from the current SQLite state.

    Steps:
    1. Back up any existing xlsx to data/backups/jobs.YYYY-MM-DD.xlsx
    2. Build a new Workbook with overview, jobs, runs, profile, and quota sheets
    3. Write atomically via atomic_xlsx_write (tmp -> rename)
    """
    xlsx_path = xlsx_path or _XLSX_PATH
    backups_dir = backups_dir or _BACKUPS_DIR

    backup_existing_xlsx(xlsx_path, backups_dir)

    wb = Workbook()
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass

    # overview sheet (default active sheet)
    ws_overview = wb.active
    _build_overview_sheet(ws_overview, conn)

    # jobs sheet
    ws_jobs = wb.create_sheet("jobs")
    _build_jobs_sheet(ws_jobs, conn)

    # runs sheet
    ws_runs = wb.create_sheet("runs")
    _build_runs_sheet(ws_runs, conn)

    # profile sheet
    ws_profile = wb.create_sheet("profile")
    _build_profile_sheet(ws_profile)

    # quota sheet
    ws_quota = wb.create_sheet("quota")
    _build_quota_sheet(ws_quota, conn)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    atomic_xlsx_write(wb, xlsx_path)
    logger.info("Workbook regenerated: %s", xlsx_path)
